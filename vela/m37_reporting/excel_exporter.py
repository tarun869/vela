"""
Excel report exporter for VPP financial and operational data.

Produces formatted Excel workbooks for settlement data, revenue
attribution, and operational KPI reports using openpyxl.
"""
from __future__ import annotations

import io
from dataclasses import dataclass
from typing import Any, Dict, List, Optional


@dataclass
class SheetSpec:
    name: str
    headers: List[str]
    rows: List[List[Any]]
    format_currency: List[int] = None   # Column indices to format as $
    format_pct: List[int] = None        # Column indices to format as %
    freeze_row: int = 1


class ExcelExporter:
    """
    Exports VPP data to formatted Excel workbooks.

    Uses openpyxl if available; falls back to CSV bytes if not.
    """

    def export(self, sheets: List[SheetSpec], title: str = "VPP Report") -> bytes:
        """
        Export sheets to an Excel workbook and return bytes.

        Returns openpyxl workbook bytes if openpyxl is installed,
        otherwise returns UTF-8 CSV of the first sheet.
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, numbers
            from openpyxl.utils import get_column_letter

            wb = openpyxl.Workbook()
            wb.remove(wb.active)   # Remove default empty sheet

            for spec in sheets:
                ws = wb.create_sheet(title=spec.name[:31])
                # Header row
                header_fill = PatternFill(fill_type="solid", fgColor="1F4E79")
                for col_idx, header in enumerate(spec.headers, 1):
                    cell = ws.cell(row=1, column=col_idx, value=header)
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center")

                # Data rows
                for row_idx, row in enumerate(spec.rows, 2):
                    for col_idx, value in enumerate(row, 1):
                        cell = ws.cell(row=row_idx, column=col_idx, value=value)
                        if spec.format_currency and (col_idx - 1) in spec.format_currency:
                            cell.number_format = '"$"#,##0.00'
                        elif spec.format_pct and (col_idx - 1) in spec.format_pct:
                            cell.number_format = "0.00%"

                # Auto-fit column widths
                for col in ws.columns:
                    max_len = max((len(str(c.value or "")) for c in col), default=8)
                    ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 40)

                # Freeze header
                ws.freeze_panes = ws.cell(row=spec.freeze_row + 1, column=1)

            buf = io.BytesIO()
            wb.save(buf)
            return buf.getvalue()

        except ImportError:
            # Fallback to CSV
            lines = []
            if sheets:
                spec = sheets[0]
                lines.append(",".join(str(h) for h in spec.headers))
                for row in spec.rows:
                    lines.append(",".join(str(v) for v in row))
            return "\n".join(lines).encode("utf-8")

    def settlement_workbook(
        self,
        settlements: List[Dict[str, Any]],
    ) -> bytes:
        """Build a settlement export workbook."""
        headers = ["Interval", "Asset ID", "Service", "Dispatched MWh",
                   "Price $/MWh", "Revenue $", "Penalties $", "Net $"]
        rows = [
            [
                s.get("interval", ""),
                s.get("asset_id", ""),
                s.get("service", ""),
                s.get("dispatched_mwh", 0.0),
                s.get("price_usd_mwh", 0.0),
                s.get("revenue_usd", 0.0),
                s.get("penalties_usd", 0.0),
                s.get("net_usd", 0.0),
            ]
            for s in settlements
        ]
        sheet = SheetSpec(
            name="Settlement",
            headers=headers,
            rows=rows,
            format_currency=[4, 5, 6, 7],
        )
        return self.export([sheet])
